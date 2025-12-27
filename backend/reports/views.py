from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from accounting.models import Voucher, JournalEntry
from inventory.models import InventoryStockItem, StockMovement
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from django.db.models import Sum, Q

class BaseExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get_workbook(self):
        wb = Workbook()
        return wb

    def save_response(self, wb, filename):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

class DayBookExcelView(BaseExcelView):
    def get(self, request):
        start_date = request.query_params.get('startDate')
        end_date = request.query_params.get('endDate')
        
        vouchers = Voucher.objects.filter(tenant_id=request.user.tenant_id)
        
        if start_date:
            vouchers = vouchers.filter(date__gte=start_date)
        if end_date:
            vouchers = vouchers.filter(date__lte=end_date)
            
        vouchers = vouchers.order_by('date', 'id')

        wb = self.get_workbook()
        ws = wb.active
        ws.title = "Day Book"

        # Headers
        headers = ["Date", "Voucher Type", "Voucher No", "Party/Ledger", "Particulars", "Amount"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        # Data
        for row_num, voucher in enumerate(vouchers, 2):
            ws.cell(row=row_num, column=1, value=voucher.date)
            ws.cell(row=row_num, column=2, value=voucher.type)
            ws.cell(row=row_num, column=3, value=voucher.voucher_number)
            
            # Party logic (simple fallback)
            party = voucher.party or "N/A"
            ws.cell(row=row_num, column=4, value=party)
            
            particulars = voucher.narration or ""
            ws.cell(row=row_num, column=5, value=particulars)
            
            # Amount logic
            amount = voucher.total or voucher.amount or 0
            ws.cell(row=row_num, column=6, value=amount)

        # Basic width adjustment
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        return self.save_response(wb, f"DayBook_{datetime.now().strftime('%Y%m%d')}.xlsx")

class LedgerReportExcelView(BaseExcelView):
    def get(self, request):
        ledger_name = request.query_params.get('ledger')
        start_date = request.query_params.get('startDate')
        end_date = request.query_params.get('endDate')
        
        if not ledger_name:
            return Response({"error": "Ledger name is required"}, status=400)

        # Logic to find vouchers involving this ledger
        # This is complex because ledger can be in party, account, from_account, to_account, or journal entries
        # For simplicity/performance in this view, we can rely on JournalEntry if it's consistent, 
        # OR query Vouchers with Q objects logic similar to Frontend.
        
        # Let's try querying JournalEntry first as it is the normalized source? 
        # But JournalEntry might not be fully populated for all historic data if migration issues existed.
        # Fallback to Voucher query matching frontend logic.
        
        vouchers = Voucher.objects.filter(tenant_id=request.user.tenant_id)
        
        if start_date:
            vouchers = vouchers.filter(date__gte=start_date)
        if end_date:
            vouchers = vouchers.filter(date__lte=end_date)
            
        # Filter by ledger involvement
        # 1. Party match
        q_party = Q(party=ledger_name)
        # 2. Account match (Payment/Receipt)
        q_account = Q(account=ledger_name)
        # 3. Contra match
        q_contra = Q(from_account=ledger_name) | Q(to_account=ledger_name)
        # 4. Journal Entry match (subquery or join)
        # Doing a simple filter on vouchers might miss journal entries if not joined properly.
        # Instead, let's fetch IDs of vouchers having journal entries for this ledger
        journal_voucher_ids = JournalEntry.objects.filter(
            tenant_id=request.user.tenant_id, 
            ledger=ledger_name
        ).values_list('voucher_id', flat=True)
        
        final_vouchers = vouchers.filter(
            q_party | q_account | q_contra | Q(id__in=journal_voucher_ids)
        ).distinct().order_by('date', 'id')

        wb = self.get_workbook()
        ws = wb.active
        ws.title = f"Ledger - {ledger_name[:20]}" # Truncate if too long

        ws.cell(row=1, column=1, value=f"Ledger Report: {ledger_name}")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        
        headers = ["Date", "Particulars", "Voucher Type", "Debit", "Credit", "Balance"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        row_num = 4
        running_balance = 0
        
        for voucher in final_vouchers:
            debit = 0
            credit = 0
            particulars = ""
            
            # Replicating Frontend Logic
            v_type = voucher.type
            
            if v_type == 'sales':
                if voucher.party == ledger_name:
                    debit = voucher.total or 0
                    particulars = "Sales"
                # Else logic (if account matches? Sales usually doesn't have account field in this model, it has party)
            
            elif v_type == 'purchase':
                if voucher.party == ledger_name:
                    credit = voucher.total or 0 # Purchase party is Creditor -> Credit? No, Purchase Party is Creditor (Liability).
                    # Wait, if I Purchase 100 from Vendor, Vendor is Credited 100. Correct.
                    particulars = "Purchases"
            
            elif v_type == 'payment':
                if voucher.party == ledger_name:
                    debit = voucher.amount or 0 # Payment to Party -> Party Debited.
                    particulars = voucher.account or "Cash"
                elif voucher.account == ledger_name:
                    credit = voucher.amount or 0 # Payment FROM Account -> Account Credited.
                    particulars = voucher.party or "Unknown"
            
            elif v_type == 'receipt':
                if voucher.party == ledger_name:
                    credit = voucher.amount or 0 # Receipt from Party -> Party Credited.
                    particulars = voucher.account or "Cash"
                elif voucher.account == ledger_name:
                    debit = voucher.amount or 0 # Receipt INTO Account -> Account Debited.
                    particulars = voucher.party or "Unknown"
            
            elif v_type == 'contra':
                if voucher.from_account == ledger_name:
                    credit = voucher.amount or 0
                    particulars = voucher.to_account
                elif voucher.to_account == ledger_name:
                    debit = voucher.amount or 0
                    particulars = voucher.from_account
            
            elif v_type == 'journal':
                 # Find specific entry for this ledger
                 entries = voucher.journal_entries.all()
                 for entry in entries:
                     if entry.ledger == ledger_name:
                         debit += entry.debit
                         credit += entry.credit
                 particulars = "Journal Entry" # Or aggregate others

            running_balance += (debit - credit)
            
            ws.cell(row=row_num, column=1, value=voucher.date)
            ws.cell(row=row_num, column=2, value=particulars)
            ws.cell(row=row_num, column=3, value=v_type)
            ws.cell(row=row_num, column=4, value=debit if debit > 0 else "")
            ws.cell(row=row_num, column=5, value=credit if credit > 0 else "")
            ws.cell(row=row_num, column=6, value=running_balance)
            row_num += 1

        return self.save_response(wb, f"Ledger_{ledger_name}_{datetime.now().strftime('%Y%m%d')}.xlsx")

class TrialBalanceExcelView(BaseExcelView):
    def get(self, request):
        # Determine balances from Journal Entries (Source of Truth)
        # Note: This requires Journal Entries to be up to date.
        
        entries = JournalEntry.objects.filter(tenant_id=request.user.tenant_id)
        
        # Aggregation
        ledger_balances = entries.values('ledger').annotate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit')
        ).order_by('ledger')
        
        wb = self.get_workbook()
        ws = wb.active
        ws.title = "Trial Balance"
        
        headers = ["Ledger", "Debit", "Credit"]
        for col_num, header in enumerate(headers, 1):
             ws.cell(row=1, column=col_num, value=header).font = Font(bold=True)
             
        row = 2
        total_debit = 0
        total_credit = 0
        
        for item in ledger_balances:
            debit = item['total_debit'] or 0
            credit = item['total_credit'] or 0
            
            # Net off provided logic requests standard TB (usually net balance per ledger)
            # Frontend logic: if debit > credit return { debit: d-c, credit: 0 }
            
            net_debit = 0
            net_credit = 0
            
            if debit > credit:
                net_debit = debit - credit
            elif credit > debit:
                net_credit = credit - debit
            
            if net_debit == 0 and net_credit == 0:
                continue
                
            ws.cell(row=row, column=1, value=item['ledger'])
            ws.cell(row=row, column=2, value=net_debit if net_debit > 0 else "")
            ws.cell(row=row, column=3, value=net_credit if net_credit > 0 else "")
            
            total_debit += net_debit
            total_credit += net_credit
            row += 1
            
        # Totals
        ws.cell(row=row, column=1, value="Total").font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_debit).font = Font(bold=True)
        ws.cell(row=row, column=3, value=total_credit).font = Font(bold=True)
        
        return self.save_response(wb, f"TrialBalance_{datetime.now().strftime('%Y%m%d')}.xlsx")

class StockSummaryExcelView(BaseExcelView):
    def get(self, request):
         # Placeholder - Inventory model logic might be needed
         # Based on StockItems and Voucher Items
         wb = self.get_workbook()
         ws = wb.active
         ws.title = "Stock Summary"
         ws.cell(row=1, column=1, value="Stock Summary Download logic not fully connected to backend inventory models yet.")
         return self.save_response(wb, "StockSummary.xlsx")

class GSTReportExcelView(BaseExcelView):
    def get(self, request):
         wb = self.get_workbook()
         ws = wb.active
         ws.title = "GST Report"
         ws.cell(row=1, column=1, value="GST Report logic placeholder.")
         return self.save_response(wb, "GSTReport.xlsx")
